#!/usr/bin/env python3
"""Extract HWS AI Club use-case data from AI_Use_Cases_by_Major_HWS.xlsx
into site/data.json and site/js/data.js for the static HWSAICLUB website.

Rerunnable: safe to run again any time the source spreadsheet changes.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "AI_Use_Cases_by_Major_HWS.xlsx"
SITE_DIR = ROOT / "site"
DATA_JSON_PATH = SITE_DIR / "data.json"
DATA_JS_PATH = SITE_DIR / "js" / "data.js"

EXPECTED_HEADER = ["#", "Use Case", "Difficulty", "Level", "Description", "HWS Program Page"]
DIFFICULTY_TO_LEVEL = {"Easy": 1, "Medium": 2, "Hard": 3}


def slugify(name):
    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return slug


def load_majors_index(wb):
    ws = wb["Majors Index"]
    majors = []
    for row in ws.iter_rows(min_row=3, max_row=44, min_col=1, max_col=3, values_only=True):
        number, name, program_link = row
        if number is None and name is None and program_link is None:
            continue
        assert number and name and program_link, f"Incomplete Majors Index row: {row}"
        majors.append({"number": number, "name": name, "programLink": program_link})
    assert len(majors) == 42, f"Expected 42 majors in index, found {len(majors)}"
    return majors


def load_use_cases(ws, sheet_name):
    header = [c.value for c in ws[2]]
    assert header == EXPECTED_HEADER, f"Unexpected header in sheet '{sheet_name}': {header}"

    rows = []
    for row in ws.iter_rows(min_row=3, max_row=22, min_col=1, max_col=6, values_only=True):
        number, title, difficulty, level, description, _program_page = row
        assert number and title, f"Empty use case row in sheet '{sheet_name}': {row}"
        assert difficulty in DIFFICULTY_TO_LEVEL, (
            f"Invalid difficulty '{difficulty}' in sheet '{sheet_name}' row {number}"
        )
        assert level == DIFFICULTY_TO_LEVEL[difficulty], (
            f"Difficulty/Level mismatch in sheet '{sheet_name}' row {number}: "
            f"difficulty={difficulty}, level={level}"
        )
        assert description, f"Empty description in sheet '{sheet_name}' row {number}"
        rows.append(
            {
                "number": number,
                "title": title.strip(),
                "difficulty": difficulty,
                "level": level,
                "description": description.strip(),
            }
        )

    assert len(rows) == 20, f"Expected 20 use cases in sheet '{sheet_name}', found {len(rows)}"

    # Confirm no stray 21st row.
    row_23 = [c.value for c in ws[23]]
    assert all(v is None for v in row_23), f"Unexpected data in row 23 of sheet '{sheet_name}': {row_23}"

    return rows


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: source workbook not found at {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    majors_index = load_majors_index(wb)

    sheet_names = [name for name in wb.sheetnames if name != "Majors Index"]
    assert len(sheet_names) == 42, f"Expected 42 major sheets, found {len(sheet_names)}"
    assert len(sheet_names) == len(majors_index)

    majors_out = []
    seen_slugs = set()
    difficulty_counts = {"Easy": 0, "Medium": 0, "Hard": 0}

    for index_entry, sheet_name in zip(majors_index, sheet_names):
        ws = wb[sheet_name]
        use_cases = load_use_cases(ws, sheet_name)

        for uc in use_cases:
            difficulty_counts[uc["difficulty"]] += 1

        slug = slugify(index_entry["name"])
        assert slug not in seen_slugs, f"Duplicate slug '{slug}' for major '{index_entry['name']}'"
        seen_slugs.add(slug)

        majors_out.append(
            {
                "number": index_entry["number"],
                "slug": slug,
                "name": index_entry["name"],
                "programLink": index_entry["programLink"],
                "useCases": use_cases,
            }
        )

    total_use_cases = sum(len(m["useCases"]) for m in majors_out)
    assert total_use_cases == 840, f"Expected 840 total use cases, found {total_use_cases}"
    assert sum(difficulty_counts.values()) == 840

    output = {"majors": majors_out}

    SITE_DIR.mkdir(exist_ok=True)
    (SITE_DIR / "js").mkdir(exist_ok=True)

    DATA_JSON_PATH.write_text(json.dumps(output, indent=2) + "\n", encoding="utf-8")
    DATA_JS_PATH.write_text(
        "window.HWS_AI_DATA = " + json.dumps(output, indent=2) + ";\n",
        encoding="utf-8",
    )

    print("HWSAICLUB data extraction complete")
    print(f"  Majors:     {len(majors_out)}")
    print(f"  Use cases:  {total_use_cases}")
    print(
        f"  Difficulty: Easy={difficulty_counts['Easy']} "
        f"Medium={difficulty_counts['Medium']} Hard={difficulty_counts['Hard']}"
    )
    print(f"  Wrote: {DATA_JSON_PATH.relative_to(ROOT)}")
    print(f"  Wrote: {DATA_JS_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
